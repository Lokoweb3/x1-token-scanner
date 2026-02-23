#!/usr/bin/env python3
"""Export burned LP tokens to Excel spreadsheet"""

import asyncio
import aiohttp
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

INCINERATOR = "1nc1nerator11111111111111111111111111111111"
RPC_URL = "https://rpc.mainnet.x1.xyz"

async def rpc_request(method: str, params: list):
    async with aiohttp.ClientSession() as session:
        payload = {"jsonrpc": "2.0", "id": 1, "method": method, "params": params}
        async with session.post(RPC_URL, json=payload) as resp:
            result = await resp.json()
            return result.get("result")

async def identify_lp(lp_mint: str) -> dict:
    """Identify what token pair an LP belongs to"""
    result = {
        "pool_address": None,
        "token_a_mint": None,
        "token_b_mint": None,
    }
    
    lp_holders = await rpc_request("getTokenLargestAccounts", [lp_mint])
    
    if not lp_holders or not lp_holders.get("value"):
        return result
    
    for holder in lp_holders["value"][:5]:
        holder_addr = holder.get("address", "")
        
        holder_info = await rpc_request("getAccountInfo", [holder_addr, {"encoding": "jsonParsed"}])
        if not holder_info or not holder_info.get("value"):
            continue
        
        data = holder_info["value"].get("data", {})
        if not isinstance(data, dict):
            continue
        
        owner = data.get("parsed", {}).get("info", {}).get("owner", "")
        
        if owner == INCINERATOR:
            continue
        
        result["pool_address"] = owner
        
        pool_tokens = await rpc_request(
            "getTokenAccountsByOwner",
            [owner, {"programId": "TokenkegQfeZyiNwAJbNbGKPFXCWuBvf9Ss623VQ5DA"}, {"encoding": "jsonParsed"}]
        )
        
        if pool_tokens and pool_tokens.get("value"):
            tokens = []
            for tok in pool_tokens["value"]:
                tok_info = tok.get("account", {}).get("data", {}).get("parsed", {}).get("info", {})
                mint = tok_info.get("mint", "")
                tokens.append(mint)
            
            if len(tokens) >= 1:
                result["token_a_mint"] = tokens[0]
            if len(tokens) >= 2:
                result["token_b_mint"] = tokens[1]
        
        break
    
    return result

async def get_lp_supply(lp_mint: str) -> tuple:
    """Get LP total supply and decimals"""
    lp_info = await rpc_request("getAccountInfo", [lp_mint, {"encoding": "jsonParsed"}])
    if lp_info and lp_info.get("value"):
        data = lp_info["value"].get("data", {})
        if isinstance(data, dict):
            parsed = data.get("parsed", {})
            if parsed.get("type") == "mint":
                info = parsed.get("info", {})
                supply = int(info.get("supply", 0))
                decimals = info.get("decimals", 9)
                return supply / (10 ** decimals), decimals
    return 0, 9

async def main():
    print("Fetching and analyzing burned LP tokens...")
    print("This may take a few minutes...\n")
    
    # Get all burned tokens
    tokens = await rpc_request(
        "getTokenAccountsByOwner",
        [INCINERATOR, {"programId": "TokenkegQfeZyiNwAJbNbGKPFXCWuBvf9Ss623VQ5DA"}, {"encoding": "jsonParsed"}]
    )
    
    if not tokens or not tokens.get("value"):
        print("No tokens found")
        return
    
    data = []
    total = len(tokens["value"])
    
    for i, token_acc in enumerate(tokens["value"], 1):
        parsed = token_acc.get("account", {}).get("data", {}).get("parsed", {})
        info = parsed.get("info", {})
        mint = info.get("mint", "")
        burned_amount = float(info.get("tokenAmount", {}).get("uiAmount", 0) or 0)
        decimals = info.get("tokenAmount", {}).get("decimals", 9)
        burn_account = token_acc.get("pubkey", "")
        
        print(f"Processing {i}/{total}: {mint[:16]}...")
        
        # Get LP total supply
        total_supply, _ = await get_lp_supply(mint)
        burn_pct = (burned_amount / total_supply * 100) if total_supply > 0 else 0
        
        # Identify the pool
        lp_info = await identify_lp(mint)
        
        data.append({
            "LP Mint": mint,
            "Burned Amount": burned_amount,
            "Total Supply": total_supply,
            "Burn %": burn_pct,
            "Pool Address": lp_info["pool_address"] or "",
            "Token A": lp_info["token_a_mint"] or "",
            "Token B": lp_info["token_b_mint"] or "",
            "Burn Account": burn_account,
        })
        
        await asyncio.sleep(0.1)
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Sort by burn percentage descending
    df = df.sort_values("Burn %", ascending=False)
    
    # Create Excel workbook with formatting
    wb = Workbook()
    ws = wb.active
    ws.title = "Burned LP Tokens"
    
    # Headers
    headers = ["#", "LP Mint", "Burned Amount", "Total Supply", "Burn %", "Pool Address", "Token A", "Token B", "Burn Account"]
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for row_idx, row_data in enumerate(df.itertuples(), 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=row_data._1)  # LP Mint
        ws.cell(row=row_idx, column=3, value=row_data._2)  # Burned Amount
        ws.cell(row=row_idx, column=4, value=row_data._3)  # Total Supply
        
        # Burn % with color coding
        burn_pct = row_data._4
        burn_cell = ws.cell(row=row_idx, column=5, value=f"{burn_pct:.2f}%")
        if burn_pct >= 99:
            burn_cell.fill = PatternFill("solid", fgColor="00FF00")  # Green
        elif burn_pct >= 50:
            burn_cell.fill = PatternFill("solid", fgColor="FFFF00")  # Yellow
        elif burn_pct > 0:
            burn_cell.fill = PatternFill("solid", fgColor="FFA500")  # Orange
        
        ws.cell(row=row_idx, column=6, value=row_data._5)  # Pool Address
        ws.cell(row=row_idx, column=7, value=row_data._6)  # Token A
        ws.cell(row=row_idx, column=8, value=row_data._7)  # Token B
        ws.cell(row=row_idx, column=9, value=row_data._8)  # Burn Account
    
    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 48
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 48
    ws.column_dimensions['G'].width = 48
    ws.column_dimensions['H'].width = 48
    ws.column_dimensions['I'].width = 48
    
    # Save
    output_file = "x1_burned_lp_tokens.xlsx"
    wb.save(output_file)
    
    print(f"\n✅ Exported to: {output_file}")
    print(f"Total LP tokens analyzed: {len(data)}")

if __name__ == "__main__":
    asyncio.run(main())
